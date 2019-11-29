import openpyxl

path = r"C:/Users/daniel.vasev/Desktop/Boohoo/Lichna/Python/TestMaterials/EmptyTestData.xlsx"


workbook = openpyxl.load_workbook(path)

sheet = workbook.active

#Creating rows and columns in the file
for r in range(1,7):
    for c in range(1,2):
     sheet.cell(row=r,column= c ).value = "Welcome"

for r in range(2,7):
    for c in range(2,6):
     sheet.cell(row=r,column= c ).value = "Petra"

# Save changes in the spreadshet
workbook.save(path)