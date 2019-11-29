#import openyxl module
import openpyxl
#Selecting file location
path = r"C:\Users\daniel.vasev\Desktop\Boohoo\Lichna\Python\TestMaterials\Data1.xlsx"

#creating variable to control the excel file
workbook = openpyxl.load_workbook(path)

#Selecting file
sheet = workbook.active   # We can specify the name of the tab by using this command:  workbook.get_sheet_by_name("Sheet1")

#Check how many rows and columns the file have
row = sheet.max_row
cols = sheet.max_column

# print how many rows and col we have in the file
print(row)
print(cols)

for r in range(1,row + 1 ):
    for c in range(1, cols + 1):
        print(sheet.cell(row = r, column=c).value, end="     ")

    print()



